#!/usr/bin/env python
# coding: utf-8

# In[11]:


import openpyxl
from random import uniform


# In[5]:


pedidos = openpyxl.load_workbook('pedidos.xlsx')
pedidos.sheetnames
planilha1 = pedidos['Página1']

print(planilha1['b4'])

for linha in planilha1:
    if linha[0].value is not None:
        print(linha[0].value, end='')
    if linha[1].value is not None:
        print(linha[1].value, end='')
    if linha[2].value is not None:
        print(linha[2].value)


# In[14]:


pedidos = openpyxl.load_workbook('pedidos.xlsx')
pedidos.sheetnames
planilha1 = pedidos['Página1']

for linha in range(5,100):
    numero_pedido = linha - 1
    planilha1.cell(linha, 1).value = numero_pedido
    planilha1.cell(linha, 2).value = 1200 + linha
    preco = round(uniform(10,500),2)
    planilha1.cell(linha, 3).value = preco

pedidos.save('nova_planilha.xlsx')


# In[ ]:




